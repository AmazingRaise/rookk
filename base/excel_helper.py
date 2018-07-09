#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2018/7/4 下午9:11
# @Author  : Aries
# @Site    : 
# @File    : excel_helper.py
# @Software: PyCharm Community Edition

import openpyxl
import time
import os
import threading


def Singleton(cls):
    _instance = {}

    def _singleton(*args, **kargs):
        if cls not in _instance:
            _instance[cls] = cls(*args, **kargs)
        return _instance[cls]

    return _singleton


@Singleton
class ExcelHelper(object):

    def __init__(self):
        self.file_name = None
        self.wb = None
        self.active_sheet = None
        self.rows = None
        self.cols = None
        self.table = None
        self.mutex = threading.Lock()

    def _set_filename(self):
        week = time.strftime("%W")
        year = time.strftime("%Y")
        file_name = year + '_' + week + '.xlsx'
        self.file_name = os.path.abspath(os.path.join('./data', file_name))

    def insert_data(self, content):
        self.mutex.acquire()
        try:
            self._set_filename()
            if os.path.isfile(self.file_name):
                self.load()
            else:
                self.new_excel()
            self.append_table(content)
            self.save_excel()
        except Exception as e:
            print(e)
        finally:
            self.mutex.release()

    def load(self):
        self.wb = openpyxl.load_workbook(self.file_name)
        self.table = self.wb.active
        self.get_rows()

    def new_excel(self):
        self.wb = openpyxl.Workbook()
        self.table = self.wb.active
        self.rows = 0
        self.cols = 0

    def append_table(self, content):
        if isinstance(content, list):
            # 追加到最后
            self.table.append(content)

    def save_excel(self):
        self.wb.save(self.file_name)

    def get_rows(self):
        self.rows = self.table.max_row   #获取行数
        self.cols = self.table.max_column    #获取列数

    def get_sheets(self):
        return self.wb.sheet_names
